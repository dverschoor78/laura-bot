"""
Domínio: Relatórios Financeiros

Gera fluxos e relatórios de pagamentos, consolidando:
  - Lançamentos (transações)
  - Documentos vinculados (NFe, Recibos, Orçamentos)
  - Itens e descrições extraídas via IA
  - Percentual de quitação por pedido

Relatórios disponíveis:
  - gerar_relatorio_pagamentos() → todos os pagamentos consolidados
  - gerar_fluxo_pagamentos_obra() → fluxo por obra com detalhes de NFe/Recibo
"""

import sqlite3
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _criar_diretorio_relatorios():
    """Garante que data/relatorios/ existe."""
    relatorios_dir = Path(__file__).parent.parent / "data" / "relatorios"
    relatorios_dir.mkdir(parents=True, exist_ok=True)
    return relatorios_dir


def _extrair_nfe(nome_arquivo: str) -> str:
    """Extrai número da NFe do nome do arquivo."""
    if not nome_arquivo:
        return "---"
    match = re.search(r'NFe\s*(\d+)', nome_arquivo, re.IGNORECASE)
    if match:
        return match.group(1)
    return "---"


def _extrair_descricao(dados_claude: str) -> str:
    """Extrai resumo dos itens comprados do dados_claude."""
    if not dados_claude:
        return "---"

    # Tenta extrair "Resumo da compra:" ou similar
    match = re.search(r'Resumo da compra[:\s]+([^\n]+)', dados_claude, re.IGNORECASE)
    if match:
        desc = match.group(1).strip()
        return desc[:80]

    # Tenta extrair primeira linha significativa
    lines = dados_claude.split('\n')
    for line in lines:
        line = line.strip()
        if line and not line.startswith('**') and len(line) > 10:
            return line[:80]

    return "---"


def _parse_data(data_str: str) -> str:
    """Normaliza data para DD/MM/YYYY."""
    if not data_str:
        return None

    # Tenta padrão DD/MM/YYYY
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', data_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}/{int(month):02d}/{year}"

    # Tenta formato "N de MM de YYYY"
    match = re.search(r'(\d{1,2})\s+de\s+(\d{1,2})\s+de\s+(\d{4})', data_str)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}/{int(month):02d}/{year}"

    return None


def _setup_estilos_excel():
    """Retorna um dicionário com estilos padrão para Excel."""
    return {
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "total_fill": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "total_font": Font(bold=True, size=11),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "wrap_align": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "border": Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def gerar_fluxo_pagamentos_obra(db_path: Path, ggv: str = None, output_dir: Path = None) -> Path:
    """
    Gera fluxo de pagamentos por obra.

    Colunas: ENTRADA, C CUSTO, PEDIDO, FORNECEDOR, CATEGORIA, DESCRICAO, VALOR, NFe/Recibo, VALOR PAGO, % QUITADO

    Args:
        db_path: Caminho do banco de dados (data/laura.db)
        ggv: Filtrar por obra específica (ex: "GGV03"). Se None, retorna todas.
        output_dir: Diretório para salvar Excel. Padrão: data/relatorios/

    Returns:
        Path do arquivo Excel gerado.
    """
    if output_dir is None:
        output_dir = _criar_diretorio_relatorios()

    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    # Query base
    query = """
        SELECT
            l.pfm_codigo,
            l.ggv,
            l.fornecedor,
            l.valor,
            l.valor_pago,
            l.data_pagamento,
            l.categoria,
            d.nome as doc_nfe,
            d.dados_claude
        FROM lancamentos l
        LEFT JOIN documentos d ON l.doc_id_nfe = d.id
        WHERE l.status = 'pago'
    """
    params = []

    if ggv:
        query += " AND l.ggv = ?"
        params.append(ggv)

    query += " ORDER BY l.data_pagamento, l.ggv"

    rows = con.execute(query, params).fetchall()

    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fluxo Pagamentos"

    estilos = _setup_estilos_excel()

    # Cabeçalho
    headers = [
        "ENTRADA",
        "C CUSTO",
        "PEDIDO",
        "FORNECEDOR",
        "CATEGORIA",
        "DESCRICAO",
        "VALOR",
        "NFe/Recibo",
        "VALOR PAGO",
        "% QUITADO"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = estilos["header_fill"]
        cell.font = estilos["header_font"]
        cell.alignment = estilos["center_align"]
        cell.border = estilos["border"]

    # Dados
    row_num = 2
    total_valor = 0
    total_pago = 0

    cat_map = {
        'material': 'MATERIAL',
        'mo': 'MO',
        'servicos': 'SERVIÇOS',
        'taxa': 'TAXA',
        'imposto': 'IMPOSTO',
    }

    for row in rows:
        data_entrada = _parse_data(row['data_pagamento'] or '')
        cc = row['ggv'] or '---'
        pedido = row['pfm_codigo'] or '---'
        fornecedor = row['fornecedor'] or '---'
        categoria = cat_map.get(row['categoria'], row['categoria'] or '---').upper()
        descricao = _extrair_descricao(row['dados_claude'])
        valor = row['valor'] or 0
        valor_pago = row['valor_pago'] or 0
        nfe = _extrair_nfe(row['doc_nfe'])
        percent_quitado = (valor_pago / valor * 100) if valor > 0 else 0

        total_valor += valor
        total_pago += valor_pago

        cells_data = [
            data_entrada,
            cc,
            pedido,
            fornecedor,
            categoria,
            descricao,
            valor,
            nfe,
            valor_pago,
            percent_quitado
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)

            if isinstance(value, (int, float)):
                if col_num == 7:  # VALOR
                    cell.value = value
                    cell.number_format = 'R$ #,##0.00'
                elif col_num == 9:  # VALOR PAGO
                    cell.value = value
                    cell.number_format = 'R$ #,##0.00'
                elif col_num == 10:  # % QUITADO
                    cell.value = value / 100
                    cell.number_format = '0.0%'
                else:
                    cell.value = value
            else:
                cell.value = value

            cell.border = estilos["border"]

            if col_num in [1, 2, 3, 10]:
                cell.alignment = estilos["center_align"]
            elif col_num == 6:  # DESCRICAO
                cell.alignment = estilos["wrap_align"]
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_num += 1

    # Linha de total
    total_row = row_num
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = estilos["total_font"]
    ws.cell(row=total_row, column=1).fill = estilos["total_fill"]
    ws.cell(row=total_row, column=1).border = estilos["border"]

    for col in range(2, 10):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = estilos["total_fill"]
        cell.border = estilos["border"]

    cell_valor = ws.cell(row=total_row, column=7)
    cell_valor.value = total_valor
    cell_valor.font = estilos["total_font"]
    cell_valor.number_format = 'R$ #,##0.00'
    cell_valor.border = estilos["border"]

    cell_pago = ws.cell(row=total_row, column=9)
    cell_pago.value = total_pago
    cell_pago.font = estilos["total_font"]
    cell_pago.number_format = 'R$ #,##0.00'
    cell_pago.border = estilos["border"]

    # Largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 12

    # Altura das linhas
    ws.row_dimensions[1].height = 25
    for i in range(2, row_num):
        ws.row_dimensions[i].height = 35

    con.close()

    # Salva
    ggv_suffix = f"_{ggv}" if ggv else ""
    output_file = output_dir / f"fluxo_pagamentos{ggv_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)

    return output_file


def gerar_relatorio_pagamentos(db_path: Path, output_dir: Path = None) -> Path:
    """
    Gera relatório consolidado de todos os pagamentos.

    Colunas: ENTRADA, C CUSTO, CATEGORIA, FONTE, PEDIDO, PFM, FORNECEDOR, CNPJ/CPF, TIPO, FORMA PGTO, PAGO

    Args:
        db_path: Caminho do banco de dados (data/laura.db)
        output_dir: Diretório para salvar Excel. Padrão: data/relatorios/

    Returns:
        Path do arquivo Excel gerado.
    """
    if output_dir is None:
        output_dir = _criar_diretorio_relatorios()

    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    rows = con.execute("""
        SELECT
            l.pfm_codigo,
            l.ggv,
            l.fornecedor,
            l.valor_pago,
            l.data_pagamento,
            l.categoria,
            l.tipo_documento,
            d.condicao_pgto,
            f.cnpj,
            f.cpf
        FROM lancamentos l
        LEFT JOIN documentos d ON l.doc_id = d.id
        LEFT JOIN fornecedores f ON LOWER(f.nome) = LOWER(l.fornecedor)
        WHERE l.status = 'pago'
        ORDER BY l.data_pagamento, l.ggv
    """).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"

    estilos = _setup_estilos_excel()

    headers = [
        "ENTRADA",
        "C CUSTO",
        "CATEGORIA",
        "FONTE DO RECURSO",
        "PEDIDO",
        "PFM",
        "FORNECEDOR/CLIENTE",
        "CNPJ/CPF",
        "TIPO",
        "FORMA PGTO",
        "PAGO"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = estilos["header_fill"]
        cell.font = estilos["header_font"]
        cell.alignment = estilos["center_align"]
        cell.border = estilos["border"]

    # Dados
    row_num = 2
    total = 0

    cat_map = {
        'material': 'MATERIAL',
        'mo': 'MO',
        'servicos': 'SERVIÇOS',
        'taxa': 'TAXA',
        'imposto': 'IMPOSTO',
    }

    for row in rows:
        data_entrada = _parse_data(row['data_pagamento'] or '')
        ggv = row['ggv'] or '---'
        categoria = cat_map.get(row['categoria'], row['categoria'] or '---').upper()
        fonte = f"VII - MP CC {ggv}" if ggv != '---' else '---'
        pedido = row['pfm_codigo'] or '---'
        tipo_doc = (row['tipo_documento'] or '---').upper() if row['tipo_documento'] else '---'
        fornecedor = (row['fornecedor'] or '---')
        cnpj = row['cnpj']
        cpf = row['cpf']
        cnpj_cpf = cnpj or cpf or '---'
        tipo_forn = 'PJ' if cnpj else 'PF' if cpf else '---'

        forma = '---'
        if row['condicao_pgto']:
            forma_lower = row['condicao_pgto'].lower()
            if 'pix' in forma_lower:
                forma = 'PIX'
            elif 'boleto' in forma_lower:
                forma = 'BOLETO'
            elif 'ted' in forma_lower:
                forma = 'TED'

        valor_pago = row['valor_pago'] or 0
        total += valor_pago

        cells_data = [
            data_entrada,
            ggv,
            categoria,
            fonte,
            pedido,
            tipo_doc,
            fornecedor[:25],
            str(cnpj_cpf)[:17],
            tipo_forn,
            forma,
            valor_pago
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)

            if isinstance(value, (int, float)):
                cell.value = value
                cell.number_format = 'R$ #,##0.00'
            else:
                cell.value = value

            cell.border = estilos["border"]

            if col_num in [1, 2, 3, 11]:
                cell.alignment = estilos["center_align"]
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_num += 1

    # Total
    total_row = row_num
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = estilos["total_font"]
    ws.cell(row=total_row, column=1).fill = estilos["total_fill"]
    ws.cell(row=total_row, column=1).border = estilos["border"]

    for col in range(2, 11):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = estilos["total_fill"]
        cell.border = estilos["border"]

    cell_total = ws.cell(row=total_row, column=11)
    cell_total.value = total
    cell_total.font = estilos["total_font"]
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.border = estilos["border"]

    # Largura
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 6
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 14

    con.close()

    output_file = output_dir / f"relatorio_pagamentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_file)

    return output_file


if __name__ == "__main__":
    # Teste local
    db = Path(__file__).parent.parent / "data" / "laura.db"
    print(f"Gerando fluxo de pagamentos...")
    arquivo = gerar_fluxo_pagamentos_obra(db)
    print(f"[OK] {arquivo.name}")

    print(f"Gerando relatorio de pagamentos...")
    arquivo = gerar_relatorio_pagamentos(db)
    print(f"[OK] {arquivo.name}")
